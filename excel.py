import openpyxl as O
import xlrd


lista_r = []
Excel_file = 'C:\\Users\\AndersonGuilherme\\Desktop\\sheets\\testepy.xlsx'
Excel_worksheet = 'Sheet1'
wb = O.load_workbook(Excel_file)
ws = wb[Excel_worksheet]

row_count = sheet.nrows
col_count = sheet.ncols





b2 = ws['B2']
b3 = ws['B3']
b4 = ws['B4']


for resultado in b2.value, b3.value, b4.value:

    lista_r.append(resultado)




